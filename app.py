import os
import pandas as pd
import streamlit as st
import random
from openpyxl import Workbook

# Функция для загрузки данных из файлов в папке "clients"
# def load_client_files(folder_path='clients'):
#     all_files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')]
#     dataframes = []
#     for file in all_files:
#         df = pd.read_excel(os.path.join(folder_path, file))
#         dataframes.append(df)
#     return dataframes

# Функция для инициализации таблицы-шаблона с заголовками и подзаголовкам
def initialize_template_table(headers, subheaders):
    template_df = pd.DataFrame({
        "Заголовки": headers,
        "Подзаголовки": subheaders,
        "Выбранные товары": [[] for _ in range(len(headers))]  # Пустые списки для товаров
    })
    return template_df

# Функция для маппинга данных в итоговую таблицу и сохранения в Excel
def save_to_excel(mapped_data, output_filename):
    wb = Workbook()
    ws = wb.active

    # Записываем данные в Excel по группам
    row_num = 1

    for category, items in mapped_data.items():
        # Вставляем название категории
        ws[f'A{row_num}'] = category
        row_num += 1

        # Вставляем заголовки
        ws[f'A{row_num}'] = "Наименование"
        ws[f'B{row_num}'] = "Цена, руб"
        ws[f'C{row_num}'] = "Количество"
        row_num += 1

        # Вставляем данные
        for item in items:
            ws[f'A{row_num}'] = item["Наименование"]
            ws[f'B{row_num}'] = item.get("Цена", "")  # Цена может быть не указана
            ws[f'C{row_num}'] = item.get("Количество", "")  # Количество может быть не указано
            row_num += 1

        # Пустая строка между категориями
        row_num += 1

    # Сохранение файла
    wb.save(output_filename)
    print(f"Файл успешно сохранен как {output_filename}")

# Загрузка данных для второй таблицы (например, текущий файл Excel)
file_path = 'Каталог_Чинт.xlsx'  # Укажи путь к файлу Excel
df2 = pd.read_excel(file_path, header=14)  # Чтение с заголовками из 15-й строки

# Очищаем заголовки второй таблицы от лишних пробелов и спецсимволов
df2.columns = df2.columns.str.strip()
df2.columns = df2.columns.str.replace(r'[\n\r\t]', ' ', regex=True)
df2.columns = df2.columns.str.replace(r'\s+', ' ', regex=True)

# Фильтруем корректные столбцы
valid_columns = [col for col in df2.columns if not col.startswith("Unnamed")]
df2 = df2[valid_columns]

# Столбцы по умолчанию
default_columns = [col for col in ["Наименование", "Тариф с НДС, руб"] if col in valid_columns]

# Загрузка данных для заголовков и подзаголовков из шаблонного файла
headers_df = pd.read_excel('2055_ ЯКНО-ВВ(ВК)-6кВ_Сотрудник.xlsx')
headers = headers_df.columns.tolist()  # Получаем список заголовков
subheaders = headers_df.iloc[0].tolist()  # Получаем подзаголовки из первой строки

# Инициализация таблицы-шаблона с заголовками и подзаголовками
template_df = initialize_template_table(headers, subheaders)

# Установка кастомных стилей для уменьшения отступов и растягивания блоков
st.markdown(
    """
    <style>
    .block-container {
        padding-top: 1.7rem;
        padding-bottom: 0rem;
        padding-left: 1rem;
        padding-right: 1rem;
    }
    .css-1p05t01 {
        padding: 0;
    }
    .st-dataframe {
        width: 100%;
    }
    </style>
    """, 
    unsafe_allow_html=True
)

# Интерфейс Streamlit
st.title("Табличный интерфейс")

# Первый блок: данные из файлов в папке clients
st.subheader("Данные клиентов")
client_data = load_client_files()  # Загрузка данных из файлов клиентов
if client_data:
    selected_client = st.selectbox("Выберите файл клиента", options=[f for f in os.listdir('clients') if f.endswith('.xlsx')])
    client_df = pd.read_excel(os.path.join('clients', selected_client))

    # Выбор необходимых столбцов для отображения в клиентских данных
    client_columns = ["Наименование", "Тариф с НДС, руб", "Выбрать"]
    if all(col in client_df.columns for col in client_columns):
        client_df_filtered = client_df[client_columns]
    else:
        client_df_filtered = client_df  # Показать все, если нужные столбцы отсутствуют

    st.dataframe(client_df_filtered, use_container_width=True, hide_index=True)  # Таблица на всю ширину контейнера
else:
    st.write("Нет данных для отображения")

# Второй блок: таблица с файла из Каталог_Чинт.xlsx с чекбоксами
st.subheader("Таблица товаров")

# Добавление поля для поиска и фильтров на одной строке
with st.container():
    col1, col2, col3 = st.columns([3, 1, 1])
    with col1:
        search_query = st.text_input("Поиск товаров", "")
    with col2:
        selected_header = st.selectbox("Раздел", headers)
    with col3:
        selected_subheader = st.selectbox("Подраздел", subheaders)

# Поиск по таблице. Если ничего не введено, отображается пустая таблица
if search_query.strip():
    filtered_df2 = df2[df2.apply(lambda row: row.astype(str).str.contains(search_query, case=False).any(), axis=1)]
    show_table = True  # Показываем таблицу, если что-то найдено
else:
    filtered_df2 = pd.DataFrame(columns=df2.columns)  # Пустая таблица
    show_table = False  # Скрываем таблицу, если нет запроса

# Используем st.data_editor для отображения данных с чекбоксами
if show_table:
    filtered_df2["Выбрать"] = False  # Добавляем колонку для чекбоксов
    edited_df2 = st.data_editor(filtered_df2[default_columns + ["Выбрать"]], num_rows="dynamic", use_container_width=True)
    
    # Перемещение выбранных товаров в шаблон
    for index, row in edited_df2.iterrows():
        if row["Выбрать"]:
            template_df.at[headers.index(selected_header), "Выбранные товары"].append(row["Наименование"])

# Третий блок: таблица-шаблон для выбранных товаров на всю ширину
st.subheader("Выбранные товары по категориям")
st.dataframe(template_df.drop(columns=["Заголовки", "Подзаголовки"]), use_container_width=True, hide_index=True)

# Четвертый блок: Сохранение в Excel
if st.button("Сохранить в Excel"):
    # Маппинг выбранных данных в итоговую таблицу
    mapped_data = {}
    for index, row in template_df.iterrows():
        header = row["Заголовки"]
        selected_items = row["Выбранные товары"]
        if selected_items:
            mapped_data[header] = [{"Наименование": item} for item in selected_items]

    # Сохранение в Excel файл
    save_to_excel(mapped_data, 'mapped_data.xlsx')
    st.success("Файл сохранен как mapped_data.xlsx!")
